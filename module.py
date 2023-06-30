import requests
import json
import pandas as pd
import os

from utils import get_token_azul

full_path = os.path.realpath(__file__)
main_folder = os.path.dirname(full_path)
cotacoes_folder = main_folder + "./Cotacoes/"

import openpyxl
from datetime import datetime, timedelta
from openpyxl.styles import PatternFill

def fazer_post_azul(url, headers, payload):
    try:
        response = requests.post(url, headers=headers, data=json.dumps(payload))
        response.raise_for_status()

        return response.json()
    except requests.exceptions.RequestException as e:
        return "Erro ao fazer requisição: " + str(e)

def obter_lowest_points(origem, destino, data_inicial, data_final):
    url = "https://b2c-api.voeazul.com.br/tudoAzulReservationAvailability/api/tudoazul/reservation/availability/v4/availability"

    headers = {
        "Accept": "application/json, text/plain, */*",
        "Accept-Encoding": "gzip, deflate, br",
        "Accept-Language": "en-ZA,en;q=0.9",
        "Authorization": f"Bearer {get_token_azul()}",
        "Content-Type": "application/json",
        "Culture": "pt-BR",
        "Device": "novosite",
        "Ocp-Apim-Subscription-Key": "0fc6ff296ef2431bb106504c92dd227c",
        "Referer": "https://www.voeazul.com.br/",
        "Sec-Ch-Ua": "\"Not.A/Brand\";v=\"8\", \"Chromium\";v=\"114\", \"Google Chrome\";v=\"114\"",
        "Sec-Ch-Ua-Mobile": "?0",
        "Sec-Ch-Ua-Platform": "\"Windows\"",
        "Sec-Fetch-Dest": "empty",
        "Sec-Fetch-Mode": "cors",
        "Sec-Fetch-Site": "same-site",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Safari/537.36"
    }

    payload = {
        "criteria": [
            {
                "departureStation": origem,
                "arrivalStation": destino,
                "std": "",  # Será substituído no loop
                "departureDate": ""  # Será substituído no loop
            }
        ],
        "passengers": [
            {
                "type": "ADT",
                "count": "1",
                "companionPass": "false"
            }
        ],
        "flexibleDays": {
            "daysToLeft": "3",
            "daysToRight": "3"
        },
        "currencyCode": "BRL"
    }

    current_date = datetime.strptime(data_inicial, "%d/%m/%Y")
    end_date = datetime.strptime(data_final, "%d/%m/%Y")
    lowest_points_list = []

    while current_date <= end_date:
        payload["criteria"][0]["std"] = current_date.strftime("%m/%d/%Y")
        payload["criteria"][0]["departureDate"] = current_date.strftime("%Y-%m-%d")

        response_data = fazer_post_azul(url, headers, payload)

        # Verificar se a resposta não é nula
        if response_data and "data" in response_data and "trips" in response_data["data"] and response_data["data"]["trips"]:
            fare_info = response_data["data"]["trips"][0].get("fareInformation")
            lowest_points = fare_info["lowestPoints"] if fare_info and "lowestPoints" in fare_info else "Voo Não Disponível"
            lowest_points_list.append({"Data": current_date.strftime("%Y-%m-%d"), "Lowest Points": lowest_points})
        else:
            lowest_points_list.append({"Data": current_date.strftime("%Y-%m-%d"), "Lowest Points": "Voo Não Disponível"})

        current_date += timedelta(days=1)

    #ajustar o formato das datas para o nome do arquivo
    data_inicial = data_inicial.replace("/", "_")
    data_final = data_final.replace("/", "_")
    data_inicial = data_inicial[:-5]
    data_final = data_final[:-5]
    cotacao_filename = "Cotacao_" + origem + destino + "_" + data_inicial + "ate" + data_final + ".xlsx"


    df = pd.DataFrame(lowest_points_list)
    df.to_excel(cotacoes_folder + cotacao_filename, index=False)

    workbook = openpyxl.load_workbook(cotacoes_folder + cotacao_filename)
    sheet = workbook.active
    min_lowest_points = min(row['Lowest Points'] for row in lowest_points_list if isinstance(row['Lowest Points'], int))

    # Aplicar formatação condicional
    for row in sheet.iter_rows(min_row=2, min_col=2, max_col=2):
        for cell in row:
            if cell.value == min_lowest_points:
                # Formatar a célula em verde
                fill = PatternFill(fill_type="solid", fgColor="00FF00")
                cell.fill = fill

    #o nome do arquivo deve conter as variaveis origem destino e datas
    workbook.save(cotacoes_folder + cotacao_filename)
    
    return lowest_points_list

